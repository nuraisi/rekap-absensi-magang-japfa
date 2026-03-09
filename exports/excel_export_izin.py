import io
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.pagebreak import Break
from openpyxl.utils import get_column_letter

bulan_indo = {
    1: "Januari",
    2: "Februari",
    3: "Maret",
    4: "April",
    5: "Mei",
    6: "Juni",
    7: "Juli",
    8: "Agustus",
    9: "September",
    10: "Oktober",
    11: "November",
    12: "Desember"
}

def format_tanggal_indo(tanggal):
    if pd.isna(tanggal):
        return ""
    return f"{tanggal.day} {bulan_indo[tanggal.month]} {tanggal.year}"

def export_excel_izin(df_izin):

    df = df_izin.copy()

    df["Tanggal Mulai Izin"] = pd.to_datetime(
        df["Tanggal Mulai Izin"],
        errors="coerce"
    )

    df["Tanggal Akhir Izin"] = pd.to_datetime(
        df["Tanggal Akhir Izin"],
        errors="coerce"
    )
    # ======================
    # WORKBOOK
    # ======================
    wb = Workbook()
    ws = wb.active
    ws.title = "Daftar Izin"

    ws.print_options.horizontalCentered = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = False
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

    # ======================
    # STYLE
    # ======================
    header_font = Font(bold=True, size=8)
    small_font = Font(size=7)

    center = Alignment(horizontal="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    grey_fill = PatternFill(
        start_color="D9D9D9",
        end_color="D9D9D9",
        fill_type="solid"
    )

    current_row = 1

    # ======================
    # LOOP PER DEPARTEMEN
    # ======================
    for i, (departemen, df_dept) in enumerate(df.groupby("Departemen")):

        # Page break antar departemen
        if i != 0:
            ws.row_breaks.append(Break(id=current_row - 1))

        df_dept = df_dept.sort_values(
            by=["Nama", "Tanggal Mulai Izin"]
        )

        # ======================
        # JUDUL
        # ======================
        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=5
        )

        cell = ws.cell(current_row, 1)
        cell.value = "DAFTAR IZIN MAHASISWA"
        cell.font = Font(size=14, bold=True)
        cell.alignment = center

        current_row += 1

        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=5
        )

        cell = ws.cell(current_row, 1)
        cell.value = f"Departemen: {departemen}"
        cell.font = Font(size=11, bold=True)
        cell.alignment = center

        current_row += 2

        # ======================
        # HEADER TABEL (HANYA SEKALI)
        # ======================
        header = [
            "No",
            "Nama",
            "Tanggal Mulai Izin",
            "Tanggal Selesai Izin",
            "Keterangan"
        ]

        for col, val in enumerate(header, start=1):

            cell = ws.cell(current_row, col)
            cell.value = val
            cell.font = header_font
            cell.alignment = center
            cell.border = thin_border
            cell.fill = grey_fill

        current_row += 1

        # ======================
        # LOOP PER NAMA
        # ======================
        for nama, df_nama in df_dept.groupby("Nama"):

            for no, (_, row) in enumerate(df_nama.iterrows(), start=1):

                ws.cell(current_row, 1, no).alignment = center
                ws.cell(current_row, 1).font = small_font

                ws.cell(current_row, 2, row["Nama"]).font = small_font

                ws.cell(
                    current_row,
                    3,
                    format_tanggal_indo(row["Tanggal Mulai Izin"])
                    if pd.notna(row["Tanggal Mulai Izin"]) else ""
                ).font = small_font

                ws.cell(
                    current_row,
                    4,
                    format_tanggal_indo(row["Tanggal Akhir Izin"])
                    if pd.notna(row["Tanggal Akhir Izin"]) else ""
                ).font = small_font

                ws.cell(
                    current_row,
                    5,
                    row.get("Alasan", "")
                ).font = small_font

                for c in range(1, 6):
                    ws.cell(current_row, c).border = thin_border

                current_row += 1

            # spasi antar nama
            current_row += 1

    # ======================
    # WIDTH KOLOM
    # ======================
    widths = [4, 20, 15, 15, 25]

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ======================
    # SAVE
    # ======================
    buffer = io.BytesIO()

    wb.save(buffer)

    buffer.seek(0)

    return buffer