import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break
from openpyxl.styles import PatternFill
from datetime import datetime, time
import streamlit as st

grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

TARIF_HARIAN = 20000
POTONGAN_TELAT = 12000
TAMBAHAN_LEMBUR = 12000
MIN_JAM_KERJA = time(4,0)
BATAS_LEMBUR = time(12,0)


def auto_column_width(ws):
    from openpyxl.utils import get_column_letter

    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)

        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col_idx).value
            if val:
                max_length = max(max_length, len(str(val)))

        ws.column_dimensions[col_letter].width = max_length + 3

def extract_date_range(filename):

    match_range = re.search(r"(\d{1,2})-(\d{1,2})", filename)
    if not match_range:
        raise ValueError("Nama file harus mengandung rentang tanggal contoh: 12-17")

    start_day = int(match_range.group(1))
    end_day = int(match_range.group(2))

    return list(range(start_day, end_day + 1))



def hitung_uang_absensi(row):

    telat = row.get("Telat", 0)
    att_time = row.get("ATT_Time")

    # pastikan telat angka
    try:
        telat = int(telat)
    except:
        telat = 0

    # ubah string ke time
    if isinstance(att_time, str):
        try:
            att_time = datetime.strptime(att_time, "%H:%M").time()
        except:
            att_time = None

    # ubah pandas timestamp
    if hasattr(att_time, "time"):
        att_time = att_time.time()

    bonus = 0
    potongan = 0

    lembur = att_time and att_time > BATAS_LEMBUR
    kurang_jam = att_time and att_time < MIN_JAM_KERJA

    # ======================
    # RULE KHUSUS
    # ======================

    # Telat + Lembur → saling hapus
    if telat == 1 and lembur:
        return 0, 0

    # Telat + <4 jam → hanya 1 potongan
    if telat == 1 and kurang_jam:
        return 0, POTONGAN_TELAT

    # ======================
    # RULE NORMAL
    # ======================

    if telat == 1:
        potongan += POTONGAN_TELAT

    if kurang_jam:
        potongan += POTONGAN_TELAT

    if lembur:
        bonus += TAMBAHAN_LEMBUR

    return bonus, potongan

def export_excel_rekap(final_absen, final_izin, filename_absen):

    def extract_bulan(filename):

        bulan_list = [
            "Januari","Februari","Maret","April","Mei","Juni",
            "Juli","Agustus","September","Oktober","November","Desember"
        ]

        for b in bulan_list:
            if b.lower() in filename.lower():
                return b

        return "Tidak diketahui"

    bulan = extract_bulan(filename_absen)

    wb = Workbook()
    ws = wb.active

    # ======================
    # SETTING PRINT FINAL
    # ======================
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = False

    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False

    ws.page_margins = PageMargins(
        left=0.2,
        right=0.2,
        top=0.3,
        bottom=0.3
    )
    ws.title = f"Rekap {bulan}"

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    small_font = Font(size=9)
    header_font = Font(size=9, bold=True)

    current_row = 1

    for departemen, df_dept in final_absen.groupby("Departemen"):

        day_cols = extract_date_range(filename_absen)

        first_day_col = 4
        last_day_col = first_day_col + len(day_cols) - 1
        jumlah_col = last_day_col + 1

        # ======================
        # JUDUL
        # ======================
        total_cols = 3 + len(
            [c for c in df_dept.columns if isinstance(c, int)]
        ) + 1  # +Jumlah

        # ======================
        # JUDUL BARIS 1
        # ======================
        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=jumlah_col
        )

        cell = ws.cell(row=current_row, column=1)
        cell.value = "DAFTAR ABSENSI MAHASISWA"
        cell.font = Font(size=14, bold=True)
        cell.alignment = Alignment(horizontal="center")
        current_row += 1

        # ======================
        # JUDUL BARIS 2 (BULAN)
        # ======================
        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=jumlah_col
        )

        cell = ws.cell(row=current_row, column=1)
        cell.value = f"BULAN {bulan.upper()}"
        cell.font = Font(size=12, bold=True)
        cell.alignment = Alignment(horizontal="center")

        current_row += 2

        # ======================
        # HEADER
        # ======================

        for d in day_cols:
            if d not in df_dept.columns:
                df_dept[d] = ""

        header = ["No Id", "Nama", "Departemen"]
        header += [str(d) for d in day_cols]
        header += ["Jumlah"]

        for col_idx, col_name in enumerate(header, start=1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = col_name
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            cell.fill = grey_fill

        ws.row_dimensions[current_row].height = 20

        current_row += 1

        start_data_row = current_row

        # ======================
        # DATA PER ORANG
        # ======================
        for _, row in df_dept.iterrows():

            # ambil nilai telat dan att_time jika ada
            row_telat = row.get("Telat", 0)
            row_att = row.get("ATT_Time", None)

            # buat row baru untuk perhitungan
            row_calc = row.copy()
            row_calc["Telat"] = row_telat
            row_calc["ATT_Time"] = row_att

            cell = ws.cell(row=current_row, column=1)
            cell.value = row["No.Akun"]
            cell.alignment = Alignment(horizontal="center")
            cell.font = small_font

            cell = ws.cell(row=current_row, column=2)
            cell.value = row["Nama"]
            cell.font = small_font

            cell = ws.cell(row=current_row, column=3)
            cell.value = row["Departemen"]
            cell.alignment = Alignment(horizontal="center")
            cell.font = small_font

            # Kolom tanggal
            for i, d in enumerate(day_cols):
                val = row.get(d, "")
                cell = ws.cell(row=current_row, column=first_day_col + i)
                cell.value = 1 if val in [1, "1", True] else ""
                cell.font = small_font
                cell.alignment = Alignment(horizontal="center")

            # ======================
            # FORMULA JUMLAH DINAMIS
            # ======================
            start_letter = get_column_letter(first_day_col)
            end_letter = get_column_letter(last_day_col)

            bonus = row.get("Bonus", 0)
            potongan = row.get("Potongan", 0)

            formula = (
                f"=SUM({start_letter}{current_row}:{end_letter}{current_row})"
                f"*{TARIF_HARIAN}"
                f"+{bonus}"
                f"-{potongan}"
            )

            jumlah_cell = ws.cell(row=current_row, column=jumlah_col)
            jumlah_cell.value = formula
            jumlah_cell.number_format = '#,##0'
            jumlah_cell.font = Font(size=9, bold=False)

            # Border seluruh baris
            for col in range(1, jumlah_col + 1):
                ws.cell(row=current_row, column=col).border = thin_border
            
            ws.row_dimensions[current_row].height = 18
            current_row += 1

        end_data_row = current_row - 1

        # ======================
        # TOTAL PER DEPARTEMEN (FORMULA)
        # ======================

        # merge semua kolom kecuali jumlah
        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=jumlah_col - 1
        )

        cell = ws.cell(row=current_row, column=1)
        cell.value = "TOTAL"
        cell.font = Font(size=9, bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = grey_fill

        # border untuk area merge
        for col in range(1, jumlah_col):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.fill = grey_fill

        # kolom jumlah
        jumlah_letter = get_column_letter(jumlah_col)

        total_formula = (
            f"=SUM({jumlah_letter}{start_data_row}:"
            f"{jumlah_letter}{end_data_row})"
        )

        total_cell = ws.cell(row=current_row, column=jumlah_col)
        total_cell.value = total_formula
        total_cell.number_format = '#,##0'
        total_cell.font = Font(size=9, bold=True)
        total_cell.border = thin_border
        total_cell.fill = grey_fill
        total_cell.alignment = Alignment(horizontal="center")
        ws.row_dimensions[current_row].height = 20

        # ======================
        # KETERANGAN
        # ======================

        current_row += 2

        # KET
        ws.merge_cells(
            start_row=current_row,
            start_column=2,
            end_row=current_row,
            end_column=jumlah_col
        )

        cell = ws.cell(row=current_row, column=2)
        cell.value = "Ket:"
        cell.font = Font(size=9, bold=True)

        current_row += 1

        ws.merge_cells(
            start_row=current_row,
            start_column=2,
            end_row=current_row,
            end_column=jumlah_col
        )

        cell = ws.cell(row=current_row, column=2)
        cell.value = "Uang Makan Rp. 12.000,- dan Transport Rp. 8.000,-"
        cell.font = Font(size=9)

        current_row += 1

        ws.merge_cells(
            start_row=current_row,
            start_column=2,
            end_row=current_row,
            end_column=jumlah_col
        )

        cell = ws.cell(row=current_row, column=2)
        cell.value = "Jika terlambat tanpa keterangan hanya mendapat Transport Rp. 8.000,-"
        cell.font = Font(size=9)

        # ======================
        # TTD
        # ======================
        bulan_id = {
            "January":"Januari","February":"Februari","March":"Maret",
            "April":"April","May":"Mei","June":"Juni",
            "July":"Juli","August":"Agustus","September":"September",
            "October":"Oktober","November":"November","December":"Desember"
        }

        today = datetime.today()
        bulan_ttd = bulan_id[today.strftime("%B")]
        today_str = f"{today.day:02d} {bulan_ttd} {today.year}"

        current_row += 2

        # kiri
        ws.merge_cells(
            start_row=current_row,
            start_column=2,
            end_row=current_row,
            end_column=5
        )

        cell = ws.cell(row=current_row, column=2)
        cell.value = "Mengetahui,\nPembimbing Lapangan"
        cell.font = Font(size=9)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        ws.row_dimensions[current_row].height = 35

        # kanan
        start_col = max(6, jumlah_col // 2 + 1)

        ws.merge_cells(
            start_row=current_row,
            start_column=start_col,
            end_row=current_row,
            end_column=jumlah_col
        )

        cell = ws.cell(row=current_row, column=start_col)
        cell.value = f"Sidoarjo, {today_str}"
        cell.font = Font(size=9)
        cell.alignment = Alignment(horizontal="right")

        # spasi tanda tangan
        current_row += 4

        start_col = max(6, jumlah_col // 2 + 1)

        ws.merge_cells(
            start_row=current_row,
            start_column=start_col,
            end_row=current_row,
            end_column=jumlah_col
        )

        cell = ws.cell(row=current_row, column=start_col)
        cell.value = "(R. Rizki Hendri Kusrini)"
        cell.font = Font(size=9)
        cell.alignment = Alignment(horizontal="right")

        current_row += 1


        last_col_letter = get_column_letter(jumlah_col)

        ws.print_area = f"A1:{last_col_letter}{current_row}"

        # ======================
        # PAGE BREAK PER DEPARTEMEN
        # ======================
        ws.row_breaks.append(Break(id=current_row))

        current_row += 2  # lanjut spasi untuk departemen berikutnya

    ws.freeze_panes = None
    # ======================
    # FIXED WIDTH - PRINT SAFE 1-31
    # ======================

    ws.column_dimensions["A"].width = 7.5
    ws.column_dimensions["B"].width = 21
    ws.column_dimensions["C"].width = 12

    # Kolom tanggal (dinamis)
    for i in range(first_day_col, last_day_col + 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = 2.5

    # Jumlah
    ws.column_dimensions[get_column_letter(jumlah_col)].width = 12

    # ======================
    # SIMPAN KE MEMORY
    # ======================
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer